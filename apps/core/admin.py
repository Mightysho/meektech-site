# from django.contrib import admin
# from django.utils.html import format_html
# from django.http import HttpResponse
# from django.shortcuts import render
# from django.urls import path
# from django.db.models import Count
# from openpyxl import Workbook
# import csv

# from .models import Visitor
# from .admin_forms import VisitorDateRangeExportForm
# from django.utils import timezone
# from django.conf import settings


# def excel_datetime(dt):
#     if not dt:
#         return None
#     if timezone.is_aware(dt):
#         return timezone.make_naive(dt)
#     return dt

# @admin.register(Visitor)
# class VisitorAdmin(admin.ModelAdmin):

#     list_display = (
#         "ip_address",
#         "country",
#         "city",
#         "path",
#         "latitude",
#         "longitude",
#         "visited_at",
#         "view_map_icon",
#     )


    
#     list_filter = ("country", "city", "visited_at")
#     search_fields = ("ip_address", "path")
#     actions = ("export_visitors_csv", "export_visitors_excel")
#     readonly_fields = ('latitude', 'longitude')


#     # ======================
#     # MAP ICON
#     # ======================
#     def view_map_icon(self, obj):
#         if not obj.latitude or not obj.longitude:
#             return "-"
            
#         return format_html(
#             '<a href="javascript:void(0);" class="map-btn" data-lat="{}" data-lng="{}" '
#             'style="cursor:pointer;" title="View location">'
#             '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" '
#             'fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">'
#             '<path d="M21 10c0 7-9 13-9 13S3 17 3 10a9 9 0 1 1 18 0z"/>'
#             '<circle cx="12" cy="10" r="3"/></svg>'
#             '</a>',
#             obj.latitude,
#             obj.longitude,
#             obj.country or "",
#             obj.city or ""
#         )
#     view_map_icon.short_description = "Map"

#     class Media:
#         js = ("js/custom-admin.js",)

#     # ======================
#     # CSV EXPORT (FILTERED / SELECTED)
#     # ======================
#     def export_visitors_csv(self, request, queryset):
#         response = HttpResponse(content_type="text/csv")
#         response["Content-Disposition"] = 'attachment; filename="meektech-site-visitors.csv"'

#         writer = csv.writer(response)
#         writer.writerow([
#             "IP Address", "Country", "City", "Path",
#             "Latitude", "Longitude", "Visited At"
#         ])

#         for v in queryset:
#             writer.writerow([
#                 v.ip_address, v.country, v.city, v.path,
#                 v.latitude, v.longitude, v.visited_at
#             ])

#         return response

#     export_visitors_csv.short_description = "Export selected visitors (CSV)"

#     # ======================
#     # EXCEL EXPORT + SUMMARY
#     # ======================
#     def export_visitors_excel(self, request, queryset):
#         wb = Workbook()

#         # -------- Sheet 1: Visitors --------
#         ws = wb.active
#         ws.title = "Visitors"
#         ws.append([
#             "IP Address", "Country", "City", "Path",
#             "Latitude", "Longitude", "Visited At"
#         ])

#         for v in queryset:
#             visited_at = (
#             timezone.make_naive(v.visited_at)
#             if v.visited_at and timezone.is_aware(v.visited_at)
#             else v.visited_at
#         )
#             ws.append([
#                 v.ip_address, v.country, v.city, v.path,
#                 v.latitude, v.longitude, excel_datetime(v.visited_at)
#             ])
#         visited_at.strftime("%Y-%m-%d %H:%M:%S") if visited_at else ""

#         # -------- Sheet 2: Summary --------
#         summary = wb.create_sheet(title="Summary")

#         total_visits = queryset.count()
#         summary.append(["Metric", "Value"])
#         summary.append(["Total Visits", total_visits])
#         summary.append([])

#         # Top Countries
#         summary.append(["Top Countries", "Visits"])
#         for row in (
#             queryset.values("country")
#             .annotate(count=Count("id"))
#             .order_by("-count")[:10]
#         ):
#             summary.append([row["country"] or "Unknown", row["count"]])

#         summary.append([])

#         # Top Cities
#         summary.append(["Top Cities", "Visits"])
#         for row in (
#             queryset.values("city")
#             .annotate(count=Count("id"))
#             .order_by("-count")[:10]
#         ):
#             summary.append([row["city"] or "Unknown", row["count"]])

#         response = HttpResponse(
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#         response["Content-Disposition"] = 'attachment; filename="meektech-site-visitors.xlsx"'
#         wb.save(response)

#         return response

#     export_visitors_excel.short_description = "Export selected visitors (Excel + Summary)"

#     # ======================
#     # DATE RANGE EXPORT VIEW
#     # ======================
#     def get_urls(self):
#         urls = super().get_urls()
#         custom_urls = [
#             path(
#                 "export-date-range/",
#                 self.admin_site.admin_view(self.export_by_date_range),
#                 name="visitor_export_date_range",
#             )
#         ]
#         return custom_urls + urls

#     def export_by_date_range(self, request):
#         form = VisitorDateRangeExportForm(request.POST or None)

#         if form.is_valid():
#             qs = Visitor.objects.filter(
#                 visited_at__date__range=(
#                     form.cleaned_data["start_date"],
#                     form.cleaned_data["end_date"],
#                 )
#             )
#             return self.export_visitors_excel(request, qs)

#         return render(
#             request,
#             "admin/core/visitor/export_date_range.html",
#             {"form": form, "title": "Export Visitors by Date Range"},
#         )


# from django.contrib.sites.models import Site

# Site._meta.verbose_name = "Site Domain ID"
# Site._meta.verbose_name_plural = "Website Domains IDs"

from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from django.db.models import Count
from openpyxl import Workbook
import csv

from .models import Visitor
from .admin_forms import VisitorDateRangeExportForm
from django.utils import timezone
from django.conf import settings


def excel_datetime(dt):
    if not dt:
        return None
    if timezone.is_aware(dt):
        return timezone.make_naive(dt)
    return dt

@admin.register(Visitor)
class VisitorAdmin(admin.ModelAdmin):

    list_display = (
        "ip_address",
        "country",
        "city",
        "path",
        "latitude",
        "longitude",
        "visited_at",
        "view_map_icon",
    )


    
    list_filter = ("country", "city", "visited_at")
    search_fields = ("ip_address", "path")
    actions = ("export_visitors_csv", "export_visitors_excel")
    # readonly_fields = ('latitude', 'longitude')


    # ======================
    # MAP ICON
    # ======================
    def view_map_icon(self, obj):
        if not obj.latitude or not obj.longitude:
            return "-"
            
        return format_html(
            '<a href="javascript:void(0);" class="map-btn" data-lat="{}" data-lng="{}" '
            'style="cursor:pointer;" title="View location">'
            '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" '
            'fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">'
            '<path d="M21 10c0 7-9 13-9 13S3 17 3 10a9 9 0 1 1 18 0z"/>'
            '<circle cx="12" cy="10" r="3"/></svg>'
            '</a>',
            obj.latitude,
            obj.longitude,
            obj.country or "",
            obj.city or ""
        )
    view_map_icon.short_description = "Map"

    class Media:
        js = ("js/custom-admin.js",)

    # ======================
    # CSV EXPORT (FILTERED / SELECTED)
    # ======================
    def export_visitors_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="meektech-site-visitors.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "IP Address", "Country", "City", "Path",
            "Latitude", "Longitude", "Visited At"
        ])

        for v in queryset:
            writer.writerow([
                v.ip_address, v.country, v.city, v.path,
                v.latitude, v.longitude, v.visited_at
            ])

        return response

    export_visitors_csv.short_description = "Export selected visitors (CSV)"

    # ======================
    # EXCEL EXPORT + SUMMARY
    # ======================
    def export_visitors_excel(self, request, queryset):
        wb = Workbook()

        # -------- Sheet 1: Visitors --------
        ws = wb.active
        ws.title = "Visitors"
        ws.append([
            "IP Address", "Country", "City", "Path",
            "Latitude", "Longitude", "Visited At"
        ])

        for v in queryset:
            visited_at = (
            timezone.make_naive(v.visited_at)
            if v.visited_at and timezone.is_aware(v.visited_at)
            else v.visited_at
        )
            ws.append([
                v.ip_address, v.country, v.city, v.path,
                v.latitude, v.longitude, excel_datetime(v.visited_at)
            ])
        visited_at.strftime("%Y-%m-%d %H:%M:%S") if visited_at else ""

        # -------- Sheet 2: Summary --------
        summary = wb.create_sheet(title="Summary")

        total_visits = queryset.count()
        summary.append(["Metric", "Value"])
        summary.append(["Total Visits", total_visits])
        summary.append([])

        # Top Countries
        summary.append(["Top Countries", "Visits"])
        for row in (
            queryset.values("country")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        ):
            summary.append([row["country"] or "Unknown", row["count"]])

        summary.append([])

        # Top Cities
        summary.append(["Top Cities", "Visits"])
        for row in (
            queryset.values("city")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        ):
            summary.append([row["city"] or "Unknown", row["count"]])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="meektech-site-visitors.xlsx"'
        wb.save(response)

        return response

    export_visitors_excel.short_description = "Export selected visitors (Excel + Summary)"

    # ======================
    # DATE RANGE EXPORT VIEW
    # ======================
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export-date-range/",
                self.admin_site.admin_view(self.export_by_date_range),
                name="visitor_export_date_range",
            )
        ]
        return custom_urls + urls

    def export_by_date_range(self, request):
        form = VisitorDateRangeExportForm(request.POST or None)

        if form.is_valid():
            qs = Visitor.objects.filter(
                visited_at__date__range=(
                    form.cleaned_data["start_date"],
                    form.cleaned_data["end_date"],
                )
            )
            return self.export_visitors_excel(request, qs)

        return render(
            request,
            "admin/core/visitor/export_date_range.html",
            {"form": form, "title": "Export Visitors by Date Range"},
        )


from django.contrib.sites.models import Site

Site._meta.verbose_name = "Site Domain ID"
Site._meta.verbose_name_plural = "Website Domains IDs"

